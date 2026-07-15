# -*- coding: utf-8 -*-
"""피드백 엑셀 생성기 (공통 도구) — 고객 폴더의 feedback/feedback_data.py를 읽어
같은 폴더에 '피드백-작성-양식.xlsx'를 생성한다.
실행: python3 common/tools/make_feedback_xlsx.py clients/<고객폴더>"""
import os
import sys

if len(sys.argv) < 2:
    sys.exit("사용법: python3 common/tools/make_feedback_xlsx.py clients/<고객폴더>")
FEEDBACK_DIR = os.path.abspath(os.path.join(sys.argv[1], "feedback"))
if not os.path.isfile(os.path.join(FEEDBACK_DIR, "feedback_data.py")):
    sys.exit(f"feedback_data.py를 찾을 수 없습니다: {FEEDBACK_DIR}")

sys.path.insert(0, FEEDBACK_DIR)
from feedback_data import (LOCATIONS, TYPES, PRIORITIES, STATUSES, MAP_OPTIONS,
                           RECEIVE_OPTIONS, FEEDBACK_ITEMS, INFO_ROWS, QUESTIONS)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

NAVY, NAVY_LIGHT, LINE, MUTED = "10294C", "EEF3FA", "E4E8F0", "5B6577"
thin = Side(style="thin", color=LINE)
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor=NAVY)
body_font = Font(name="맑은 고딕", size=10)
muted_font = Font(name="맑은 고딕", size=10, color=MUTED)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center")

MAX_ROW = 120  # 드롭다운이 걸리는 마지막 행

wb = Workbook()

# ---------- 시트1: 작성 안내 ----------
ws = wb.active
ws.title = "작성 안내"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 95
GUIDE = [
    ("title", "웹사이트 피드백 작성 안내"), ("gap", ""),
    ("body", "안녕하세요. 웹사이트 초안 검토를 부탁드립니다."),
    ("body", "PC와 휴대폰 양쪽에서 한 번씩 봐주시면 가장 좋습니다."), ("gap", ""),
    ("head", "작성 방법 (3가지만 지켜 주세요)"),
    ("body", "① [피드백 작성] 시트에 한 줄에 하나씩 적어 주세요. 위치·유형·중요도는 셀을 클릭하면 목록에서 고를 수 있습니다."),
    ("body", "② 문구 수정은 바꿀 문장을 \"그대로\" 적어 주세요.  예) 제목을 \"○○○\"로 교체"),
    ("body", "③ 디자인 수정은 참고 이미지나 참고 사이트 주소를 '참고 자료' 칸에 함께 적어 주세요."), ("gap", ""),
    ("head", "[질문] 시트에 꼭 답변해 주세요"),
    ("body", "지난 피드백을 정리하다 확인이 필요한 것들을 [질문 (답변 요청)] 시트에 모았습니다. 파란 칸에 답변을 적어 주세요."),
    ("body", "'진행 상태' 열은 작업자가 관리합니다. '확인 필요'로 표시된 항목이 질문과 연결되어 있습니다."), ("gap", ""),
    ("head", "꼭 확인해 주세요"),
    ("body", "[확정 정보] 시트는 피드백과 별개로, 실제 값을 주셔야 사이트를 오픈할 수 있는 항목입니다."), ("gap", ""),
    ("muted", "작성 후 이 파일을 그대로 회신해 주시면 됩니다. 감사합니다."),
]
r = 2
for kind, text in GUIDE:
    c = ws.cell(row=r, column=2, value=text)
    c.font = {"title": Font(name="맑은 고딕", bold=True, size=16, color=NAVY),
              "head": Font(name="맑은 고딕", bold=True, size=12, color=NAVY),
              "muted": muted_font}.get(kind, body_font)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    r += 1

# ---------- 시트2: 피드백 작성 ----------
ws2 = wb.create_sheet("피드백 작성")
headers = ["번호", "위치 (선택)", "유형 (선택)", "요청 내용 (구체적으로)",
           "참고 자료 (링크·첨부명)", "중요도 (선택)", "진행 상태 (작업자)"]
widths = [6, 24, 22, 60, 30, 16, 22]
for i, (h, w) in enumerate(zip(headers, widths), start=1):
    cell = ws2.cell(row=1, column=i, value=h)
    cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, center, border
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.row_dimensions[1].height = 24
ws2.freeze_panes = "A2"

STATUS_FILLS = {
    "완료": PatternFill("solid", fgColor="E6F4EA"),
    "진행 중": PatternFill("solid", fgColor="FFF4D6"),
    "확인 필요 (고객 답변 대기)": PatternFill("solid", fgColor="FDE8E8"),
}

question_by_feedback = {fb_no: q_no for q_no, fb_no, _ in QUESTIONS}

for idx, item in enumerate(FEEDBACK_ITEMS, start=1):
    row = idx + 1
    values = [idx, *item]
    if idx in question_by_feedback:
        ref = values[4]  # 참고 자료 열 (0:번호 1:위치 2:유형 3:요청 4:참고 5:중요도)
        tag = f"→ [질문] 시트 {question_by_feedback[idx]} 답변 요청"
        values[4] = f"{ref}\n{tag}" if ref else tag
    for col, v in enumerate(values, start=1):
        c = ws2.cell(row=row, column=col, value=v)
        c.font, c.border = body_font, border
        c.alignment = wrap if col in (4, 5) else Alignment(wrap_text=True, vertical="center",
                                                           horizontal="center")
    status = item[5]
    if status in STATUS_FILLS:
        ws2.cell(row=row, column=7).fill = STATUS_FILLS[status]

first_empty = len(FEEDBACK_ITEMS) + 2
for r in range(first_empty, MAX_ROW + 1):
    ws2.cell(row=r, column=1, value=r - 1).font = muted_font
    for col in range(1, 8):
        c = ws2.cell(row=r, column=col)
        c.border, c.font = border, body_font
        c.alignment = wrap if col in (4, 5) else center
        if r % 2 == 0:
            c.fill = PatternFill("solid", fgColor="F7F9FC")

# ---------- 숨김 시트: 드롭다운 목록 ----------
ws_list = wb.create_sheet("목록")
for col, values in enumerate([LOCATIONS, TYPES, PRIORITIES, MAP_OPTIONS,
                              RECEIVE_OPTIONS, STATUSES], start=1):
    for i, v in enumerate(values, start=1):
        ws_list.cell(row=i, column=col, value=v)
ws_list.sheet_state = "hidden"

def dv(col_letter, count, prompt):
    d = DataValidation(type="list", formula1=f"목록!${col_letter}$1:${col_letter}${count}",
                       allow_blank=True, showDropDown=False)
    d.prompt, d.showInputMessage = prompt, True
    d.error, d.showErrorMessage = "목록에서 선택해 주세요.", True
    return d

dv_loc = dv("A", len(LOCATIONS), "셀을 클릭하면 위치 목록이 나옵니다.")
dv_type = dv("B", len(TYPES), "피드백 종류를 골라 주세요.")
dv_pri = dv("C", len(PRIORITIES), "중요도를 골라 주세요.")
dv_status = dv("F", len(STATUSES), "작업자가 관리하는 칸입니다.")
for d in (dv_loc, dv_type, dv_pri, dv_status):
    ws2.add_data_validation(d)
dv_loc.add(f"B2:B{MAX_ROW}")
dv_type.add(f"C2:C{MAX_ROW}")
dv_pri.add(f"F2:F{MAX_ROW}")
dv_status.add(f"G2:G{MAX_ROW}")

# ---------- 시트3: 질문 (답변 요청) ----------
ws_q = wb.create_sheet("질문 (답변 요청)", index=2)
ws_q.sheet_view.showGridLines = False
for i, (h, w) in enumerate(zip(["번호", "관련 피드백", "질문", "답변 (여기에 작성)"],
                               [8, 12, 62, 45]), start=1):
    cell = ws_q.cell(row=1, column=i, value=h)
    cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, center, border
    ws_q.column_dimensions[get_column_letter(i)].width = w
ws_q.row_dimensions[1].height = 24
ws_q.freeze_panes = "A2"

for i, (q_no, fb_no, question) in enumerate(QUESTIONS, start=2):
    cells = [
        ws_q.cell(row=i, column=1, value=q_no),
        ws_q.cell(row=i, column=2, value=f"{fb_no}번 항목"),
        ws_q.cell(row=i, column=3, value=question),
        ws_q.cell(row=i, column=4, value=""),
    ]
    cells[0].font = Font(name="맑은 고딕", size=10, bold=True, color=NAVY)
    cells[1].font, cells[2].font, cells[3].font = muted_font, body_font, body_font
    cells[3].fill = PatternFill("solid", fgColor=NAVY_LIGHT)
    for c in cells:
        c.border = border
        c.alignment = wrap if c.column >= 3 else center
    ws_q.row_dimensions[i].height = max(48, 15 * (question.count("\n") + 1))

# ---------- 시트4: 확정 정보 ----------
ws3 = wb.create_sheet("확정 정보")
ws3.sheet_view.showGridLines = False
for i, (h, w) in enumerate(zip(["항목", "현재 상태 (임시값)", "확정 값 (여기에 작성)"],
                               [26, 40, 45]), start=1):
    cell = ws3.cell(row=1, column=i, value=h)
    cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, center, border
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.row_dimensions[1].height = 24
ws3.freeze_panes = "A2"

for i, (name, cur) in enumerate(INFO_ROWS, start=2):
    a = ws3.cell(row=i, column=1, value=name)
    b = ws3.cell(row=i, column=2, value=cur)
    c = ws3.cell(row=i, column=3, value="")
    a.font = Font(name="맑은 고딕", size=10, bold=True, color=NAVY)
    b.font, c.font = muted_font, body_font
    c.fill = PatternFill("solid", fgColor=NAVY_LIGHT)
    for cell in (a, b, c):
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws3.row_dimensions[i].height = 22

map_row = next(i for i, r in enumerate(INFO_ROWS, start=2) if r[0] == "지도 사용 여부")
recv_row = next(i for i, r in enumerate(INFO_ROWS, start=2) if r[0] == "상담 신청 접수 방법")
dv_map = dv("D", len(MAP_OPTIONS), "지도 사용 여부를 골라 주세요.")
dv_recv = dv("E", len(RECEIVE_OPTIONS), "상담 접수를 어떻게 받으실지 골라 주세요.")
ws3.add_data_validation(dv_map)
ws3.add_data_validation(dv_recv)
dv_map.add(f"C{map_row}")
dv_recv.add(f"C{recv_row}")

note_row = len(INFO_ROWS) + 3
note = ws3.cell(row=note_row, column=1,
                value="※ 로고 파일과 대표 사진은 파일을 메일·카톡으로 함께 보내 주세요. (PNG/JPG, 가능한 원본 크기)")
note.font = muted_font
ws3.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=3)

out = os.path.join(FEEDBACK_DIR, "피드백-작성-양식.xlsx")
wb.save(out)
print("저장 완료:", out, f"(피드백 {len(FEEDBACK_ITEMS)}건 반영)")
